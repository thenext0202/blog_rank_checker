"""설정 엑셀 파일 초기 생성 스크립트 (1회만 실행)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# 스타일 정의
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, headers, col_widths):
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# ── 시트1: 제품정보 ──
ws1 = wb.active
ws1.title = "제품정보"
headers1 = ["제품명", "제품설명", "소구점1", "소구점2", "소구점3", "소구점4", "소구점5"]
widths1 = [20, 40, 30, 30, 30, 30, 30]
style_header(ws1, headers1, widths1)

# 예시 데이터
ws1.append(["헬리컷", "헬리코박터 유산균 제품", "헬리코박터균 억제", "위 점막 보호", "장 건강 개선", "식약처 인증", ""])
ws1.append(["예시제품B", "혈압 건강 보조제", "혈압 조절 도움", "혈관 건강", "천연 원료", "", ""])

# ── 시트2: 원고유형규칙 ──
ws2 = wb.create_sheet("원고유형규칙")
headers2 = ["원고유형", "규칙1", "규칙2", "규칙3", "규칙4", "규칙5"]
widths2 = [15, 40, 40, 40, 40, 40]
style_header(ws2, headers2, widths2)

ws2.append(["후기형", "목차가 포함되어야 합니다", "실제 사용 후기가 포함되어야 합니다", "상단에 제품명이 들어가야 합니다", "소제목이 포함되어야 합니다", ""])
ws2.append(["정보형", "목차가 필요하지 않습니다", "소제목이 반드시 포함되어야 합니다", "정확한 의학/건강 정보 기반이어야 합니다", "중간에 광고 이미지 삽입 위치 표시", ""])

# ── 시트3: 공통규칙 ──
ws3 = wb.create_sheet("공통규칙")
headers3 = ["번호", "규칙내용"]
widths3 = [8, 80]
style_header(ws3, headers3, widths3)

ws3.append([1, "소제목이 꼭 들어가야 합니다"])
ws3.append([2, "중간에 광고 이미지 삽입 위치를 표시해야 합니다"])
ws3.append([3, "상단에 지정 문구가 포함되어야 합니다"])
ws3.append([4, "맞춤법과 띄어쓰기를 정확히 해야 합니다"])
ws3.append([5, "의학 정보는 사실에 기반해야 합니다"])

wb.save("C:/Users/iamhy/Desktop/프로그램 개발/manuscript_checker/검수설정.xlsx")
print("검수설정.xlsx 생성 완료!")
