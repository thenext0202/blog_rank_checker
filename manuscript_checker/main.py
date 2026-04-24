"""
원고 검수 프로그램
- 워드(.docx) 원고를 읽어 키워드 빈도, 구조, 의학 정보 정확성 등을 검수합니다.
- 설정은 검수설정.xlsx에서 관리합니다.
"""
import sys
import os
import re
import threading
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

from docx import Document
import openpyxl
import anthropic


# ─────────────────────────────────────────────
# 유틸리티
# ─────────────────────────────────────────────
def base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def resource_path(filename):
    return os.path.join(base_dir(), filename)


# ─────────────────────────────────────────────
# 엑셀 설정 읽기
# ─────────────────────────────────────────────
def load_config(path):
    """검수설정.xlsx에서 제품정보, 원고유형규칙, 공통규칙을 읽어온다."""
    config = {"products": {}, "article_types": {}, "common_rules": []}

    if not os.path.exists(path):
        return config

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # 제품정보
    if "제품정보" in wb.sheetnames:
        ws = wb["제품정보"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            name = str(row[0]).strip()
            desc = str(row[1]).strip() if row[1] else ""
            points = [str(c).strip() for c in row[2:] if c]
            config["products"][name] = {
                "description": desc,
                "selling_points": points
            }

    # 원고유형규칙
    if "원고유형규칙" in wb.sheetnames:
        ws = wb["원고유형규칙"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            atype = str(row[0]).strip()
            rules = [str(c).strip() for c in row[1:] if c]
            config["article_types"][atype] = rules

    # 공통규칙
    if "공통규칙" in wb.sheetnames:
        ws = wb["공통규칙"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                config["common_rules"].append(str(row[1]).strip())

    wb.close()
    return config


# ─────────────────────────────────────────────
# 워드 파일 읽기
# ─────────────────────────────────────────────
def read_docx(filepath):
    """워드 파일에서 텍스트와 구조 정보를 추출한다."""
    doc = Document(filepath)
    full_text = []
    paragraphs_info = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        is_heading = para.style.name.startswith('Heading')
        is_bold = any(run.bold for run in para.runs if run.text.strip())

        paragraphs_info.append({
            "text": text,
            "is_heading": is_heading,
            "is_bold": is_bold,
        })
        full_text.append(text)

    return "\n".join(full_text), paragraphs_info


# ─────────────────────────────────────────────
# 기본 분석 (로컬)
# ─────────────────────────────────────────────
def basic_analysis(text, paragraphs_info, keywords, product_info=None,
                    article_type="", type_rules=None, common_rules=None):
    """글자 수, 키워드 빈도, 소제목 여부, 규칙 준수 여부 등 기본 분석"""
    results = []
    type_rules = type_rules or []
    common_rules = common_rules or []

    # 글자 수 (공백 제외 / 포함)
    char_no_space = len(text.replace(" ", "").replace("\n", ""))
    char_with_space = len(text)
    results.append(f"[글자 수] 공백 제외: {char_no_space}자 / 공백 포함: {char_with_space}자")

    # 문장 수
    sentences = re.split(r'[.!?。]\s*', text)
    sentences = [s for s in sentences if s.strip()]
    results.append(f"[문장 수] {len(sentences)}개")

    # 키워드 빈도
    if keywords:
        results.append("")
        results.append("── 키워드 빈도 ──")
        for kw in keywords:
            kw = kw.strip()
            if not kw:
                continue
            count = text.count(kw)
            results.append(f"  '{kw}': {count}회")

    # 소제목 존재 여부
    headings = [p for p in paragraphs_info if p["is_heading"] or p["is_bold"]]
    results.append("")
    results.append(f"[소제목/강조] {len(headings)}개 발견")
    for h in headings:
        tag = "제목" if h["is_heading"] else "볼드"
        results.append(f"  [{tag}] {h['text'][:50]}")

    # ── 규칙 자동 체크 ──
    text_lower = text.lower()
    has_headings = len(headings) > 0

    # 규칙 키워드 매핑: 규칙 문장에 특정 단어가 있으면 자동 체크
    def check_rule(rule):
        """규칙 문장을 분석해서 원고에서 자동 체크할 수 있는 항목을 확인한다."""
        rule_lower = rule.lower()

        # 목차 관련
        if "목차" in rule_lower:
            if "필요하지 않" in rule_lower or "불필요" in rule_lower:
                return "SKIP", "목차 불필요 (체크 불요)"
            has_toc = any(
                "목차" in p["text"] for p in paragraphs_info
            )
            if has_toc:
                return "PASS", "목차가 포함되어 있습니다"
            else:
                return "FAIL", "목차가 발견되지 않았습니다"

        # 소제목 관련
        if "소제목" in rule_lower:
            if has_headings:
                return "PASS", f"소제목/강조 {len(headings)}개 발견"
            else:
                return "FAIL", "소제목(제목 스타일 또는 볼드)이 발견되지 않았습니다"

        # 제품명 관련
        if "제품명" in rule_lower and product_info:
            pname = product_info.get("name", "")
            if pname and pname in text:
                return "PASS", f"제품명 '{pname}'이(가) 포함되어 있습니다"
            elif pname:
                return "FAIL", f"제품명 '{pname}'이(가) 발견되지 않았습니다"

        # 이미지 삽입 위치 관련
        if "이미지" in rule_lower and ("삽입" in rule_lower or "위치" in rule_lower):
            image_markers = ["이미지", "사진", "[이미지]", "[[이미지]]", "<이미지>", "img", "광고"]
            found = any(m in text_lower for m in image_markers)
            if found:
                return "PASS", "이미지 관련 표시가 포함되어 있습니다"
            else:
                return "WARN", "이미지 삽입 위치 표시가 발견되지 않았습니다"

        # 상단 문구 관련
        if "상단" in rule_lower and ("문구" in rule_lower or "포함" in rule_lower):
            return "INFO", "상단 문구 포함 여부는 AI 검수에서 확인합니다"

        # 후기/경험 관련
        if "후기" in rule_lower or "사용 경험" in rule_lower or "사용경험" in rule_lower:
            experience_words = ["사용해", "먹어", "복용", "써보", "체험", "후기", "경험", "느낌", "효과를 봤", "달라졌"]
            found = any(w in text for w in experience_words)
            if found:
                return "PASS", "사용 경험/후기 관련 내용이 포함되어 있습니다"
            else:
                return "WARN", "사용 후기/경험 관련 표현이 부족해 보입니다"

        # 맞춤법 관련
        if "맞춤법" in rule_lower or "띄어쓰기" in rule_lower:
            return "INFO", "맞춤법/띄어쓰기는 AI 검수에서 확인합니다"

        # 의학 정보 관련
        if "의학" in rule_lower or "건강 정보" in rule_lower or "사실" in rule_lower:
            return "INFO", "의학/건강 정보 정확성은 AI 검수에서 확인합니다"

        # 자동 체크 불가 항목
        return "INFO", "AI 검수에서 확인합니다"

    # 원고 유형별 규칙 체크
    if type_rules:
        results.append("")
        results.append(f"── 원고 유형별 규칙 체크 [{article_type}] ──")
        for rule in type_rules:
            status, msg = check_rule(rule)
            if status == "PASS":
                icon = "[O 통과]"
            elif status == "FAIL":
                icon = "[X 미충족]"
            elif status == "WARN":
                icon = "[! 주의]"
            elif status == "SKIP":
                icon = "[- 해당없음]"
            else:
                icon = "[? 확인필요]"
            results.append(f"  {icon} {rule}")
            results.append(f"         -> {msg}")

    # 공통 규칙 체크
    if common_rules:
        results.append("")
        results.append("── 공통 규칙 체크 ──")
        for rule in common_rules:
            status, msg = check_rule(rule)
            if status == "PASS":
                icon = "[O 통과]"
            elif status == "FAIL":
                icon = "[X 미충족]"
            elif status == "WARN":
                icon = "[! 주의]"
            elif status == "SKIP":
                icon = "[- 해당없음]"
            else:
                icon = "[? 확인필요]"
            results.append(f"  {icon} {rule}")
            results.append(f"         -> {msg}")

    # 제품 소구점 포함 여부
    if product_info and product_info.get("selling_points"):
        results.append("")
        results.append(f"── 제품 소구점 포함 여부 [{product_info.get('name', '')}] ──")
        for sp in product_info["selling_points"]:
            if sp in text:
                results.append(f"  [O] '{sp}' - 포함됨")
            else:
                # 부분 매칭 시도 (소구점의 핵심 단어가 들어있는지)
                core_words = [w for w in sp.split() if len(w) >= 2]
                partial = any(w in text for w in core_words)
                if partial:
                    results.append(f"  [~] '{sp}' - 유사 표현 있음 (정확한 문구는 아님)")
                else:
                    results.append(f"  [X] '{sp}' - 미포함")

    return "\n".join(results)


# ─────────────────────────────────────────────
# AI 검수 (Claude API)
# ─────────────────────────────────────────────
def ai_review(api_key, text, product_info, article_type, type_rules, common_rules, keywords):
    """Claude API를 사용해 원고를 종합 검수한다."""

    # 프롬프트 구성
    prompt_parts = []
    prompt_parts.append("당신은 건강/의학 마케팅 원고를 검수하는 전문 에디터입니다.")
    prompt_parts.append("아래 원고를 꼼꼼히 검수하고, 각 항목별로 구체적인 피드백을 주세요.")
    prompt_parts.append("")

    # 제품 정보
    if product_info:
        prompt_parts.append("=" * 40)
        prompt_parts.append("[검수 대상 제품 정보]")
        prompt_parts.append(f"제품명: {product_info.get('name', '')}")
        prompt_parts.append(f"제품 설명: {product_info.get('description', '')}")
        prompt_parts.append(f"핵심 소구점: {', '.join(product_info.get('selling_points', []))}")
        prompt_parts.append("")

    # 원고 유형 및 규칙
    if article_type:
        prompt_parts.append(f"[원고 유형] {article_type}")
    if type_rules:
        prompt_parts.append(f"[유형별 규칙] {' / '.join(type_rules)}")
    if common_rules:
        prompt_parts.append(f"[공통 규칙] {' / '.join(common_rules)}")
    if keywords:
        prompt_parts.append(f"[지정 키워드] {', '.join(keywords)}")
    prompt_parts.append("")

    prompt_parts.append("=" * 40)
    prompt_parts.append("[원고 내용]")
    prompt_parts.append(text)
    prompt_parts.append("=" * 40)
    prompt_parts.append("")

    prompt_parts.append("""다음 항목을 순서대로 검수해 주세요:

1. **맞춤법/문법 검사**
   - 맞춤법, 띄어쓰기, 문법 오류를 찾아 수정안을 제시하세요.

2. **의학/건강 정보 사실 확인**
   - 원고에 포함된 의학/건강 정보가 사실에 부합하는지 확인하세요.
   - 예: 특정 수치(혈압, 혈당 등)에 대한 증상 설명이 맞는지, 성분 효능이 정확한지.
   - 잘못된 정보가 있으면 구체적으로 지적하고 올바른 정보를 제공하세요.

3. **제품 소구점 반영 여부**
   - 위에 제시된 핵심 소구점이 원고에 잘 녹아들어 있는지 확인하세요.
   - 빠진 소구점이 있으면 알려주세요.

4. **원고 유형별 규칙 준수**
   - 해당 원고 유형의 규칙을 모두 지키고 있는지 확인하세요.
   - 지키지 않은 규칙이 있으면 구체적으로 알려주세요.

5. **공통 규칙 준수**
   - 공통 가이드라인을 모두 지키고 있는지 확인하세요.

6. **가독성 및 문체**
   - 문장이 너무 길거나 어렵지 않은지, 블로그 글로서 자연스러운지 평가하세요.

7. **종합 평가**
   - 전체적인 원고 품질을 A/B/C/D 등급으로 평가하세요.
   - 핵심 개선 사항 3가지를 요약해 주세요.

각 항목별로 명확하게 구분해서 작성해 주세요.""")

    system_prompt = "\n".join(prompt_parts)

    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        messages=[
            {"role": "user", "content": system_prompt}
        ]
    )

    return message.content[0].text


# ─────────────────────────────────────────────
# GUI 애플리케이션
# ─────────────────────────────────────────────
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("원고 검수 프로그램")
        self.root.geometry("1000x750")
        self.root.minsize(800, 600)

        self.config_path = resource_path("검수설정.xlsx")
        self.config = load_config(self.config_path)
        self.api_key = self._load_api_key()
        self.current_text = ""
        self.current_paragraphs = []
        self.result_text = ""

        self._build_ui()

    # ── API 키 저장/불러오기 ──
    def _load_api_key(self):
        path = resource_path(".api_key")
        if os.path.exists(path):
            with open(path, "r") as f:
                return f.read().strip()
        return ""

    def _save_api_key(self, key):
        with open(resource_path(".api_key"), "w") as f:
            f.write(key)
        self.api_key = key

    # ── UI 구성 ──
    def _build_ui(self):
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 탭1: 원고 검수
        self.tab_review = ttk.Frame(notebook)
        notebook.add(self.tab_review, text="  원고 검수  ")
        self._build_review_tab()

        # 탭2: 설정
        self.tab_settings = ttk.Frame(notebook)
        notebook.add(self.tab_settings, text="  설정  ")
        self._build_settings_tab()

    def _build_review_tab(self):
        frame = self.tab_review

        # ── 상단: 입력 영역 ──
        input_frame = ttk.LabelFrame(frame, text="검수 설정", padding=10)
        input_frame.pack(fill=tk.X, padx=10, pady=(10, 5))

        # 행1: 파일 선택
        row1 = ttk.Frame(input_frame)
        row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="원고 파일:", width=10).pack(side=tk.LEFT)
        self.var_filepath = tk.StringVar()
        ttk.Entry(row1, textvariable=self.var_filepath, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(row1, text="파일 선택", command=self._select_file).pack(side=tk.RIGHT)

        # 행2: 제품 선택 + 원고 유형
        row2 = ttk.Frame(input_frame)
        row2.pack(fill=tk.X, pady=2)

        ttk.Label(row2, text="제품:", width=10).pack(side=tk.LEFT)
        self.var_product = tk.StringVar()
        product_names = list(self.config["products"].keys())
        self.combo_product = ttk.Combobox(row2, textvariable=self.var_product,
                                          values=product_names, state="readonly", width=20)
        self.combo_product.pack(side=tk.LEFT, padx=(0, 20))
        if product_names:
            self.combo_product.current(0)

        ttk.Label(row2, text="원고 유형:").pack(side=tk.LEFT)
        self.var_article_type = tk.StringVar()
        type_names = list(self.config["article_types"].keys())
        self.combo_type = ttk.Combobox(row2, textvariable=self.var_article_type,
                                       values=type_names, state="readonly", width=15)
        self.combo_type.pack(side=tk.LEFT, padx=(5, 0))
        if type_names:
            self.combo_type.current(0)

        # 행3: 키워드 입력
        row3 = ttk.Frame(input_frame)
        row3.pack(fill=tk.X, pady=2)
        ttk.Label(row3, text="키워드:", width=10).pack(side=tk.LEFT)
        self.var_keywords = tk.StringVar()
        ttk.Entry(row3, textvariable=self.var_keywords).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(row3, text="(쉼표로 구분)", foreground="gray").pack(side=tk.LEFT, padx=5)

        # 행4: 버튼
        row4 = ttk.Frame(input_frame)
        row4.pack(fill=tk.X, pady=(8, 0))

        self.btn_basic = ttk.Button(row4, text="기본 분석", command=self._run_basic)
        self.btn_basic.pack(side=tk.LEFT, padx=(0, 5))

        self.btn_ai = ttk.Button(row4, text="AI 검수 (Claude)", command=self._run_ai_review)
        self.btn_ai.pack(side=tk.LEFT, padx=(0, 5))

        self.btn_all = ttk.Button(row4, text="전체 검수 (기본 + AI)", command=self._run_full_review)
        self.btn_all.pack(side=tk.LEFT, padx=(0, 5))

        self.btn_save = ttk.Button(row4, text="결과 저장", command=self._save_result, state=tk.DISABLED)
        self.btn_save.pack(side=tk.RIGHT)

        self.lbl_status = ttk.Label(row4, text="", foreground="blue")
        self.lbl_status.pack(side=tk.RIGHT, padx=10)

        # ── 하단: 결과 영역 ──
        result_frame = ttk.LabelFrame(frame, text="검수 결과", padding=10)
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))

        self.txt_result = scrolledtext.ScrolledText(result_frame, wrap=tk.WORD, font=("맑은 고딕", 10))
        self.txt_result.pack(fill=tk.BOTH, expand=True)

    def _build_settings_tab(self):
        frame = self.tab_settings

        # API 키 설정
        api_frame = ttk.LabelFrame(frame, text="Claude API 키", padding=10)
        api_frame.pack(fill=tk.X, padx=10, pady=10)

        row = ttk.Frame(api_frame)
        row.pack(fill=tk.X)
        ttk.Label(row, text="API 키:").pack(side=tk.LEFT)
        self.var_api_key = tk.StringVar(value=self.api_key)
        self.entry_api_key = ttk.Entry(row, textvariable=self.var_api_key, show="*", width=60)
        self.entry_api_key.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(row, text="저장", command=self._save_api_key_clicked).pack(side=tk.LEFT)

        # 설정 파일 안내
        info_frame = ttk.LabelFrame(frame, text="검수 설정 파일", padding=10)
        info_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

        ttk.Label(info_frame, text=f"설정 파일 위치: {self.config_path}").pack(anchor=tk.W)
        ttk.Label(info_frame, text="엑셀 파일을 직접 편집하여 제품 정보, 원고 유형 규칙, 공통 규칙을 관리하세요.").pack(anchor=tk.W, pady=(5, 0))
        ttk.Label(info_frame, text="시트1: 제품정보 | 시트2: 원고유형규칙 | 시트3: 공통규칙", foreground="gray").pack(anchor=tk.W)

        btn_row = ttk.Frame(info_frame)
        btn_row.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(btn_row, text="설정 파일 열기", command=self._open_config_file).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_row, text="설정 다시 불러오기", command=self._reload_config).pack(side=tk.LEFT)

        # 현재 설정 미리보기
        preview_frame = ttk.LabelFrame(frame, text="현재 로드된 설정", padding=10)
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        self.txt_config_preview = scrolledtext.ScrolledText(preview_frame, wrap=tk.WORD,
                                                            font=("맑은 고딕", 9), height=15)
        self.txt_config_preview.pack(fill=tk.BOTH, expand=True)
        self._update_config_preview()

    # ── 이벤트 핸들러 ──
    def _select_file(self):
        path = filedialog.askopenfilename(
            title="원고 파일 선택",
            filetypes=[("Word 파일", "*.docx"), ("모든 파일", "*.*")]
        )
        if path:
            self.var_filepath.set(path)
            try:
                self.current_text, self.current_paragraphs = read_docx(path)
                self.lbl_status.config(text=f"파일 로드 완료 ({len(self.current_text)}자)")
            except Exception as e:
                messagebox.showerror("오류", f"파일 읽기 실패:\n{e}")

    def _get_keywords(self):
        raw = self.var_keywords.get()
        return [k.strip() for k in raw.split(",") if k.strip()]

    def _get_product_info(self):
        name = self.var_product.get()
        if name and name in self.config["products"]:
            info = self.config["products"][name].copy()
            info["name"] = name
            return info
        return None

    def _get_type_rules(self):
        atype = self.var_article_type.get()
        if atype and atype in self.config["article_types"]:
            return self.config["article_types"][atype]
        return []

    def _display_result(self, text):
        self.txt_result.config(state=tk.NORMAL)
        self.txt_result.delete("1.0", tk.END)
        self.txt_result.insert(tk.END, text)
        self.result_text = text
        self.btn_save.config(state=tk.NORMAL)

    def _run_basic(self):
        if not self.current_text:
            messagebox.showwarning("알림", "먼저 원고 파일을 선택하세요.")
            return
        keywords = self._get_keywords()
        result = basic_analysis(
            self.current_text, self.current_paragraphs, keywords,
            product_info=self._get_product_info(),
            article_type=self.var_article_type.get(),
            type_rules=self._get_type_rules(),
            common_rules=self.config["common_rules"]
        )

        header = f"{'=' * 50}\n  기본 분석 결과\n  {datetime.now().strftime('%Y-%m-%d %H:%M')}\n{'=' * 50}\n\n"
        self._display_result(header + result)
        self.lbl_status.config(text="기본 분석 완료")

    def _run_ai_review(self):
        if not self.current_text:
            messagebox.showwarning("알림", "먼저 원고 파일을 선택하세요.")
            return
        if not self.api_key:
            messagebox.showwarning("알림", "설정 탭에서 Claude API 키를 입력하세요.")
            return

        self.lbl_status.config(text="AI 검수 중... (30초~1분 소요)")
        self.btn_ai.config(state=tk.DISABLED)
        self.btn_all.config(state=tk.DISABLED)
        self.root.update()

        def worker():
            try:
                result = ai_review(
                    api_key=self.api_key,
                    text=self.current_text,
                    product_info=self._get_product_info(),
                    article_type=self.var_article_type.get(),
                    type_rules=self._get_type_rules(),
                    common_rules=self.config["common_rules"],
                    keywords=self._get_keywords()
                )
                header = f"{'=' * 50}\n  AI 검수 결과 (Claude)\n  {datetime.now().strftime('%Y-%m-%d %H:%M')}\n{'=' * 50}\n\n"
                self.root.after(0, lambda: self._display_result(header + result))
                self.root.after(0, lambda: self.lbl_status.config(text="AI 검수 완료"))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("AI 검수 오류", str(e)))
                self.root.after(0, lambda: self.lbl_status.config(text="AI 검수 실패"))
            finally:
                self.root.after(0, lambda: self.btn_ai.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.btn_all.config(state=tk.NORMAL))

        threading.Thread(target=worker, daemon=True).start()

    def _run_full_review(self):
        if not self.current_text:
            messagebox.showwarning("알림", "먼저 원고 파일을 선택하세요.")
            return
        if not self.api_key:
            messagebox.showwarning("알림", "설정 탭에서 Claude API 키를 입력하세요.")
            return

        self.lbl_status.config(text="전체 검수 중...")
        self.btn_ai.config(state=tk.DISABLED)
        self.btn_all.config(state=tk.DISABLED)
        self.root.update()

        # 기본 분석 먼저
        keywords = self._get_keywords()
        basic_result = basic_analysis(
            self.current_text, self.current_paragraphs, keywords,
            product_info=self._get_product_info(),
            article_type=self.var_article_type.get(),
            type_rules=self._get_type_rules(),
            common_rules=self.config["common_rules"]
        )

        def worker():
            try:
                ai_result = ai_review(
                    api_key=self.api_key,
                    text=self.current_text,
                    product_info=self._get_product_info(),
                    article_type=self.var_article_type.get(),
                    type_rules=self._get_type_rules(),
                    common_rules=self.config["common_rules"],
                    keywords=keywords
                )
                header = f"{'=' * 50}\n  전체 검수 결과\n  파일: {os.path.basename(self.var_filepath.get())}\n  {datetime.now().strftime('%Y-%m-%d %H:%M')}\n{'=' * 50}\n\n"
                divider = f"\n\n{'─' * 50}\n  AI 검수 결과 (Claude)\n{'─' * 50}\n\n"

                full = header + basic_result + divider + ai_result
                self.root.after(0, lambda: self._display_result(full))
                self.root.after(0, lambda: self.lbl_status.config(text="전체 검수 완료"))
            except Exception as e:
                # 기본 분석 결과라도 보여주기
                header = f"{'=' * 50}\n  기본 분석 결과 (AI 검수 실패)\n{'=' * 50}\n\n"
                self.root.after(0, lambda: self._display_result(header + basic_result + f"\n\n[AI 오류] {e}"))
                self.root.after(0, lambda: self.lbl_status.config(text="AI 검수 실패 (기본 분석만 완료)"))
            finally:
                self.root.after(0, lambda: self.btn_ai.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.btn_all.config(state=tk.NORMAL))

        threading.Thread(target=worker, daemon=True).start()

    def _save_result(self):
        if not self.result_text:
            return

        # 기본 저장 파일명: 원고파일명_검수결과_날짜.txt
        orig_name = os.path.splitext(os.path.basename(self.var_filepath.get()))[0] if self.var_filepath.get() else "검수결과"
        default_name = f"{orig_name}_검수결과_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"

        path = filedialog.asksaveasfilename(
            title="검수 결과 저장",
            defaultextension=".txt",
            initialfile=default_name,
            filetypes=[("텍스트 파일", "*.txt"), ("모든 파일", "*.*")]
        )
        if path:
            with open(path, "w", encoding="utf-8") as f:
                f.write(self.result_text)
            self.lbl_status.config(text=f"저장 완료: {os.path.basename(path)}")

    def _save_api_key_clicked(self):
        key = self.var_api_key.get().strip()
        if key:
            self._save_api_key(key)
            messagebox.showinfo("저장", "API 키가 저장되었습니다.")
        else:
            messagebox.showwarning("알림", "API 키를 입력하세요.")

    def _open_config_file(self):
        if os.path.exists(self.config_path):
            os.startfile(self.config_path)
        else:
            messagebox.showwarning("알림", f"설정 파일이 없습니다.\n{self.config_path}")

    def _reload_config(self):
        self.config = load_config(self.config_path)

        # 콤보박스 갱신
        product_names = list(self.config["products"].keys())
        self.combo_product["values"] = product_names
        if product_names:
            self.combo_product.current(0)

        type_names = list(self.config["article_types"].keys())
        self.combo_type["values"] = type_names
        if type_names:
            self.combo_type.current(0)

        self._update_config_preview()
        self.lbl_status.config(text="설정 다시 불러오기 완료")

    def _update_config_preview(self):
        self.txt_config_preview.config(state=tk.NORMAL)
        self.txt_config_preview.delete("1.0", tk.END)

        lines = []
        lines.append("── 제품 정보 ──")
        for name, info in self.config["products"].items():
            lines.append(f"  [{name}] {info['description']}")
            lines.append(f"    소구점: {', '.join(info['selling_points'])}")
        lines.append("")

        lines.append("── 원고 유형별 규칙 ──")
        for atype, rules in self.config["article_types"].items():
            lines.append(f"  [{atype}]")
            for r in rules:
                lines.append(f"    - {r}")
        lines.append("")

        lines.append("── 공통 규칙 ──")
        for r in self.config["common_rules"]:
            lines.append(f"  - {r}")

        if not any([self.config["products"], self.config["article_types"], self.config["common_rules"]]):
            lines.append("(설정이 비어있습니다. 검수설정.xlsx 파일을 확인하세요.)")

        self.txt_config_preview.insert(tk.END, "\n".join(lines))

    # ── API 키 ──
    def _save_api_key(self, key):
        with open(resource_path(".api_key"), "w") as f:
            f.write(key)
        self.api_key = key


# ─────────────────────────────────────────────
# 실행
# ─────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
